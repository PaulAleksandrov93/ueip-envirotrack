"""
Функции представлений для управления параметрами окружающей среды, комнатами, зданиями,
ответственными лицами, измерительными приборами и аутентификацией пользователей.
"""

import logging
from django.http import HttpResponseServerError
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework.decorators import api_view
from io import BytesIO
from openpyxl import Workbook
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from datetime import datetime, timedelta
from django.contrib.auth.models import User

from backend.models import Responsible, Room, EnviromentalParameters, MeasurementInstrument, ParameterSet, ExtendedParameterSet
from .serializers import EnvironmentalParametersSerializer, RoomSelectSerializer, ResponsibleSerializer, MeasurementInstrumentSerializer, \
                        ParameterSetSerializer, Building, BuildingEnviromentalParameters, BuildingParameterSetSerializer, \
                        BuildingParameterSet, BuildingEnvironmentalParametersSerializer, ExtendedParameterSetSerializer, BuildingSerializer, RoomSerializer, AdditionalParameters, \
                        ParameterSetForStorage, ParameterSetForStorageSerializer
import logging

logger = logging.getLogger(__name__)

class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)

        # Add custom claims
        token['username'] = user.username
        # ...

        return token


class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getRoutes(request):
    """
    Возвращает список доступных маршрутов.

    Возвращает:
        Response: JSON-ответ с перечнем доступных маршрутов.
    """
    routes = [
        '/api/token',
        '/api/token/refresh',
    ]
    return Response(routes)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_current_user(request):
#     """
#     Получает информацию о текущем аутентифицированном пользователе.

#     Args:
#         request (Request): Объект HTTP-запроса.

#     Returns:
#         Response: JSON-ответ, содержащий информацию о текущем пользователе.
#     """
#     user = request.user
#     if user.is_authenticated:
#         try:
#             responsible = Responsible.objects.get(user=user)
#             serializer = ResponsibleSerializer(responsible)
#             return Response(serializer.data)
#         except Responsible.DoesNotExist:
#             return Response({'error': 'Responsible not found'}, status=404)
#     else:
#         return Response({'error': 'User not authenticated'}, status=401)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    """
    Получает информацию о текущем аутентифицированном пользователе.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий информацию о текущем пользователе.
    """
    user = request.user
    if user.is_authenticated:
        try:
            logger.info(f'Запрос на получение информации о текущем пользователе: {user.username}')
            responsible = Responsible.objects.get(user=user)
            serializer = ResponsibleSerializer(responsible)
            logger.info(f'Информация о пользователе {user.username} успешно получена')
            return Response(serializer.data)
        except Responsible.DoesNotExist:
            logger.warning(f'Ответственный для пользователя {user.username} не найден')
            return Response({'error': 'Responsible not found'}, status=404)
        except Exception as e:
            logger.error(f'Произошла ошибка при получении информации о текущем пользователе {user.username}: {e}', exc_info=True)
            return Response({'error': 'Внутренняя ошибка сервера'}, status=500)
    else:
        logger.warning('Пользователь не аутентифицирован')
        return Response({'error': 'User not authenticated'}, status=401)


# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def getResponsibles(request):
#     responsibles = Responsible.objects.all()
#     serializer = ResponsibleSerializer(responsibles, many=True)
#     return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getResponsibles(request):
    """
    Получает список всех ответственных.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий список всех ответственных.
    """
    try:
        logger.info('Запрос на получение списка всех ответственных')
        responsibles = Responsible.objects.all()
        serializer = ResponsibleSerializer(responsibles, many=True)
        logger.info('Список всех ответственных успешно получен')
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении списка всех ответственных: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def getRooms(request):
#     rooms = Room.objects.all()
#     serializer = RoomSelectSerializer(rooms, many=True)
#     return Response(serializer.data)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getRooms(request):
    """
    Получает список всех комнат.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий список всех комнат.
    """
    try:
        logger.info('Запрос на получение списка всех комнат')
        rooms = Room.objects.all()
        serializer = RoomSelectSerializer(rooms, many=True)
        logger.info('Список всех комнат успешно получен')
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении списка всех комнат: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def getRoom(request, pk):
#     try:
#         room = Room.objects.get(id=pk, has_additional_parameters=True)
#         additional_parameters = room.additional_parameters
#         serializer = RoomSelectSerializer(room)
#         data = serializer.data
#         if additional_parameters:
#             data['additional_parameters'] = {
#                 'voltage_min': additional_parameters.voltage_min,
#                 'voltage_max': additional_parameters.voltage_max,
#                 'frequency_min': additional_parameters.frequency_min,
#                 'frequency_max': additional_parameters.frequency_max,
#                 'radiation_min': additional_parameters.radiation_min,
#                 'radiation_max': additional_parameters.radiation_max,
#             }
#         return Response(data)
#     except Room.DoesNotExist:
#         return Response({'error': 'Room not found'}, status=404)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getRoom(request, pk):
    """
    Получает информацию о комнате с дополнительными параметрами по её идентификатору.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Идентификатор комнаты.

    Returns:
        Response: JSON-ответ, содержащий информацию о комнате и её дополнительных параметрах.
    """
    try:
        logger.info(f'Запрос на получение информации о комнате с id={pk}')
        room = Room.objects.get(id=pk, has_additional_parameters=True)
        additional_parameters = room.additional_parameters
        serializer = RoomSelectSerializer(room)
        data = serializer.data
        if additional_parameters:
            data['additional_parameters'] = {
                'voltage_min': additional_parameters.voltage_min,
                'voltage_max': additional_parameters.voltage_max,
                'frequency_min': additional_parameters.frequency_min,
                'frequency_max': additional_parameters.frequency_max,
                'radiation_min': additional_parameters.radiation_min,
                'radiation_max': additional_parameters.radiation_max,
            }
        logger.info(f'Информация о комнате с id={pk} успешно получена')
        return Response(data)
    except Room.DoesNotExist:
        logger.warning(f'Комната с id={pk} не найдена')
        return Response({'error': 'Room not found'}, status=404)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении информации о комнате с id={pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def getBuildings(request):
#     buildings = Building.objects.all()
#     serializer = BuildingSerializer(buildings, many=True)
#     return Response(serializer.data)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getBuildings(request):
    """
    Получает список всех зданий.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий список всех зданий.
    """
    try:
        logger.info('Запрос на получение списка всех зданий')
        buildings = Building.objects.all()
        serializer = BuildingSerializer(buildings, many=True)
        logger.info('Список всех зданий успешно получен')
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении списка зданий: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET', 'POST'])
# # @permission_classes([IsAuthenticated])
# def measurement_instrument_type_list(request):
#     if request.method == 'GET':
#         measurement_instrument_types = MeasurementInstrument.objects.all()
#         serializer = MeasurementInstrumentSerializer(measurement_instrument_types, many=True)
#         return Response(serializer.data)
#     elif request.method == 'POST':
#         serializer = MeasurementInstrumentSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
# @permission_classes([IsAuthenticated])
def measurement_instrument_type_list(request):
    """
    Обрабатывает запросы на получение списка типов измерительных инструментов и создание нового типа.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий список типов инструментов или созданный тип инструмента.
    """
    if request.method == 'GET':
        try:
            logger.info('Запрос на получение списка типов измерительных инструментов')
            measurement_instrument_types = MeasurementInstrument.objects.all()
            serializer = MeasurementInstrumentSerializer(measurement_instrument_types, many=True)
            logger.info('Список типов измерительных инструментов успешно получен')
            return Response(serializer.data)
        except Exception as e:
            logger.error(f'Произошла ошибка при получении списка типов измерительных инструментов: {e}', exc_info=True)
            return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    elif request.method == 'POST':
        try:
            logger.info('Запрос на создание нового типа измерительного инструмента')
            serializer = MeasurementInstrumentSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                logger.info('Новый тип измерительного инструмента успешно создан')
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                logger.warning('Запрос на создание нового типа измерительного инструмента не прошел валидацию')
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f'Произошла ошибка при создании нового типа измерительного инструмента: {e}', exc_info=True)
            return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET', 'PUT', 'DELETE'])
# @permission_classes([IsAuthenticated])
# def measurement_instrument_type_detail(request, pk):
#     try:
#         measurement_instrument_type = MeasurementInstrument.objects.get(pk=pk)
#     except MeasurementInstrument.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     if request.method == 'GET':
#         serializer = MeasurementInstrumentSerializer(measurement_instrument_type)
#         return Response(serializer.data)
#     elif request.method == 'PUT':
#         serializer = MeasurementInstrumentSerializer(measurement_instrument_type, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#     elif request.method == 'DELETE':
#         measurement_instrument_type.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def measurement_instrument_type_detail(request, pk):
    """
    Обрабатывает запросы на получение, обновление и удаление конкретного типа измерительного инструмента.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ типа измерительного инструмента.

    Returns:
        Response: JSON-ответ, содержащий информацию о типе инструмента, результат обновления или статус удаления.
    """
    try:
        measurement_instrument_type = MeasurementInstrument.objects.get(pk=pk)
        logger.info(f'Запрос на получение типа измерительного инструмента с id={pk}')
    except MeasurementInstrument.DoesNotExist:
        logger.error(f'Тип измерительного инструмента с id={pk} не найден')
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = MeasurementInstrumentSerializer(measurement_instrument_type)
        logger.info(f'Тип измерительного инструмента с id={pk} успешно получен')
        return Response(serializer.data)
    elif request.method == 'PUT':
        serializer = MeasurementInstrumentSerializer(measurement_instrument_type, data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'Тип измерительного инструмента с id={pk} успешно обновлен')
            return Response(serializer.data)
        else:
            logger.warning(f'Неудачная попытка обновления типа измерительного инструмента с id={pk}')
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'DELETE':
        measurement_instrument_type.delete()
        logger.info(f'Тип измерительного инструмента с id={pk} успешно удален')
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getEnviromentalParameters(request):
    try:
        user = request.user
        responsible = request.query_params.get('responsible')
        room = request.query_params.get('room')
        date = request.query_params.get('date')

        parameters = EnviromentalParameters.objects.all().prefetch_related('room', 'responsible')

        if responsible:
            parameters = parameters.filter(responsible=responsible)

        if room:
            parameters = parameters.filter(room=room)

        if date:
            created_start = datetime.strptime(date, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            created_end = created_start + timedelta(days=1)
            parameters = parameters.filter(created_at__range=(created_start, created_end))

        parameters = parameters.order_by('-created_at')

        serialized_data = []
        for parameter in parameters:
            parameter_data = EnvironmentalParametersSerializer(parameter, context={'request': request}).data
            measurement_instruments_data = MeasurementInstrumentSerializer(parameter.measurement_instruments.all(), many=True).data
            parameter_data['measurement_instruments'] = measurement_instruments_data
            serialized_data.append(parameter_data)

        return Response(serialized_data)

    except Exception as e:
        logger.error(f'Произошла ошибка во время выполнения getEnviromentalParameters: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
    
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getEnviromentalParameter(request, pk):
#     """
#     Возвращает конкретную запись с параметрами окружающей среды.

#     Args:
#         request (Request): Объект HTTP-запроса.

#     Returns:
#         Response: JSON-ответ с параметрами окружающей среды.
#     """
#     parameters = EnviromentalParameters.objects.get(id=pk)
#     serializer = EnvironmentalParametersSerializer(parameters, many=False)
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getEnviromentalParameter(request, pk):
    """
    Возвращает конкретную запись с параметрами окружающей среды.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ записи с параметрами окружающей среды.

    Returns:
        Response: JSON-ответ с параметрами окружающей среды.
    """
    try:
        parameters = EnviromentalParameters.objects.get(id=pk)
        logger.info(f'Запрос на получение параметров окружающей среды с id={pk}')
        serializer = EnvironmentalParametersSerializer(parameters, many=False)
        return Response(serializer.data)
    except EnviromentalParameters.DoesNotExist:
        logger.error(f'Параметры окружающей среды с id={pk} не найдены')
        return Response({'error': 'Enviromental parameters not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f'Произошла ошибка во время выполнения getEnviromentalParameter: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def createEnvironmentalParameters(request):
#     try:
#         # print("Request Data:", request.data)
#         room_data = request.data.get('room')
#         # print("Room Data:", room_data)
#         room = Room.objects.get(room_number=room_data.get('room_number'))
#         # print("Room:", room)
#         created_at = request.data.get('created_at')
#         existing_parameters = EnviromentalParameters.objects.filter(room=room, created_at=created_at)
#         if existing_parameters.exists():
#             print('Вы уже создавали запись на указанную дату')
#             return Response({'error': 'An entry for this room and date already exists'}, status=status.HTTP_400_BAD_REQUEST)
#         responsible_data = request.data.get('responsible')
#         # print("Responsible Data:", responsible_data)
#         responsible, _ = Responsible.objects.get_or_create(
#             first_name=responsible_data.get('first_name'),
#             last_name=responsible_data.get('last_name'),
#             patronymic=responsible_data.get('patronymic')
#         )
#         # print("Responsible:", responsible)
#         measurement_instruments_data = request.data.get('measurement_instruments')
#         # print("Measurement Instruments Data:", measurement_instruments_data)
#         measurement_instruments_data = request.data.get('measurement_instruments', [])
#         measurement_instruments = []
#         for instrument_data in measurement_instruments_data:
#             instrument_dict = {
#                 'name': instrument_data.get('name'),  
#                 'type': instrument_data.get('type'),
#                 'serial_number': instrument_data.get('serial_number'),
#                 'calibration_date': instrument_data.get('calibration_date'),
#                 'calibration_interval': instrument_data.get('calibration_interval')
#             }
#             measurement_instruments.append(instrument_dict)
#         # print("Measurement Instruments:", measurement_instruments)
#         # print('Room has additional parameters:', room.has_additional_parameters)
#         parameter_sets_data = request.data.get('parameter_sets', [])
#         extended_parameter_sets_data = request.data.get('extended_parameter_sets', [])
#         parameter_sets_for_storage_data = request.data.get('parameter_sets_for_storage', [])

#         if room.has_additional_parameters:
#             # Если есть дополнительные параметры, то используем ExtendedParameterSet
#             parameter_sets_data = []  # Оставляем parameter_sets пустым
#         elif room.is_storage:
#             parameter_sets_data = []  # Оставляем parameter_sets пустым
#             extended_parameter_sets_ids = []  # Создаем пустой список для хранения идентификаторов расширенных параметрсетов
#         else:
#             # Если нет дополнительных параметров, то используем ParameterSet
#             extended_parameter_sets_ids = []  # Создаем пустой список для хранения идентификаторов расширенных параметрсетов
#         # print("Parameter Sets data:", parameter_sets_data)
#         # print("Extended Parameter Sets data:", extended_parameter_sets_data)
#         # print("Parameter Sets for Storage data:", parameter_sets_for_storage_data)
#         parameter_set_ids = []
#         for param_set_data in parameter_sets_data:
#             parameter_set_id = param_set_data.get('id')

#             if parameter_set_id:
#                 try:
#                     parameter_set = ParameterSet.objects.get(id=parameter_set_id)
#                     parameter_set_ids.append(parameter_set.id)
#                 except ParameterSet.DoesNotExist:
#                     print(f"ParameterSet with id {parameter_set_id} does not exist")
#                     return Response({'error': f'ParameterSet with id {parameter_set_id} does not exist'}, status=status.HTTP_400_BAD_REQUEST)
#             else:
#                 serializer = ParameterSetSerializer(data=param_set_data)
#                 if serializer.is_valid():
#                     parameter_set = serializer.save()
#                     parameter_set_ids.append(parameter_set.id)
#                 else:
#                     print("ParameterSet Serializer Errors:", serializer.errors)
#                     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#         extended_parameter_sets_ids = []

#         for extended_param_set_data in extended_parameter_sets_data:
#             # Убираем поле 'id' из данных перед сериализацией
#             extended_param_set_data.pop('id', None)
#             serializer = ExtendedParameterSetSerializer(data=extended_param_set_data)
#             if serializer.is_valid():
#                 extended_parameter_set = serializer.save()
#                 extended_parameter_sets_ids.append(extended_parameter_set.id)  # Добавляем идентификатор расширенного параметрсета в список
#             else:
#                 print("ExtendedParameterSet Serializer Errors:", serializer.errors)
#                 return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#         parameter_sets_for_storage_ids = []

#         for param_set_data in parameter_sets_for_storage_data:
#             parameter_set_id = param_set_data.get('id')

#             if parameter_set_id:
#                 try:
#                     parameter_set = ParameterSetForStorage.objects.get(id=parameter_set_id)
#                     parameter_sets_for_storage_ids.append(parameter_set.id)
#                 except ParameterSetForStorage.DoesNotExist:
#                     print(f"ParameterSetForStorage with id {parameter_set_id} does not exist")
#                     return Response({'error': f'ParameterSetForStorage with id {parameter_set_id} does not exist'}, status=status.HTTP_400_BAD_REQUEST)
#             else:
#                 serializer = ParameterSetForStorageSerializer(data=param_set_data)
#                 if serializer.is_valid():
#                     parameter_set = serializer.save()
#                     parameter_sets_for_storage_ids.append(parameter_set.id)
#                 else:
#                     print("ParameterSetForStorage Serializer Errors:", serializer.errors)
#                     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#         data = {
#             'room': room_data,
#             'responsible': responsible_data,
#             'measurement_instruments': measurement_instruments,
#             'parameter_sets': parameter_sets_data,
#             'extended_parameter_sets': extended_parameter_sets_data,
#             'parameter_sets_for_storage': parameter_sets_for_storage_data,
#             'created_at': request.data.get('created_at')
#         }
#         # print("Data:", data)
#         serializer = EnvironmentalParametersSerializer(data=data, context={'request': request})
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         else:
#             print("Serializer Errors:", serializer.errors)
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#     except Room.DoesNotExist:
#         print("Room not found")
#         return Response({'error': 'Room not found'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createEnvironmentalParameters(request):
    try:
        logger.info("Получен запрос на создание параметров окружающей среды")
        room_data = request.data.get('room')
        room = Room.objects.get(room_number=room_data.get('room_number'))
        created_at = request.data.get('created_at')
        
        existing_parameters = EnviromentalParameters.objects.filter(room=room, created_at=created_at)
        if existing_parameters.exists():
            logger.warning('Запись на указанную дату уже существует')
            return Response({'error': 'An entry for this room and date already exists'}, status=status.HTTP_400_BAD_REQUEST)

        responsible_data = request.data.get('responsible')
        responsible, _ = Responsible.objects.get_or_create(
            first_name=responsible_data.get('first_name'),
            last_name=responsible_data.get('last_name'),
            patronymic=responsible_data.get('patronymic')
        )
        
        measurement_instruments_data = request.data.get('measurement_instruments', [])
        measurement_instruments = []
        for instrument_data in measurement_instruments_data:
            instrument_dict = {
                'name': instrument_data.get('name'),
                'type': instrument_data.get('type'),
                'serial_number': instrument_data.get('serial_number'),
                'calibration_date': instrument_data.get('calibration_date'),
                'calibration_interval': instrument_data.get('calibration_interval')
            }
            measurement_instruments.append(instrument_dict)
        
        parameter_sets_data = request.data.get('parameter_sets', [])
        extended_parameter_sets_data = request.data.get('extended_parameter_sets', [])
        parameter_sets_for_storage_data = request.data.get('parameter_sets_for_storage', [])

        if room.has_additional_parameters:
            parameter_sets_data = [] 
        elif room.is_storage:
            parameter_sets_data = []
            extended_parameter_sets_ids = []
        else:
            extended_parameter_sets_ids = []

        parameter_set_ids = []
        for param_set_data in parameter_sets_data:
            parameter_set_id = param_set_data.get('id')

            if parameter_set_id:
                try:
                    parameter_set = ParameterSet.objects.get(id=parameter_set_id)
                    parameter_set_ids.append(parameter_set.id)
                except ParameterSet.DoesNotExist:
                    logger.error(f"ParameterSet with id {parameter_set_id} does not exist")
                    return Response({'error': f'ParameterSet with id {parameter_set_id} does not exist'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                serializer = ParameterSetSerializer(data=param_set_data)
                if serializer.is_valid():
                    parameter_set = serializer.save()
                    parameter_set_ids.append(parameter_set.id)
                else:
                    logger.error("ParameterSet Serializer Errors: %s", serializer.errors)
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        extended_parameter_sets_ids = []
        for extended_param_set_data in extended_parameter_sets_data:
            extended_param_set_data.pop('id', None)
            serializer = ExtendedParameterSetSerializer(data=extended_param_set_data)
            if serializer.is_valid():
                extended_parameter_set = serializer.save()
                extended_parameter_sets_ids.append(extended_parameter_set.id)
            else:
                logger.error("ExtendedParameterSet Serializer Errors: %s", serializer.errors)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        parameter_sets_for_storage_ids = []
        for param_set_data in parameter_sets_for_storage_data:
            parameter_set_id = param_set_data.get('id')

            if parameter_set_id:
                try:
                    parameter_set = ParameterSetForStorage.objects.get(id=parameter_set_id)
                    parameter_sets_for_storage_ids.append(parameter_set.id)
                except ParameterSetForStorage.DoesNotExist:
                    logger.error(f"ParameterSetForStorage with id {parameter_set_id} does not exist")
                    return Response({'error': f'ParameterSetForStorage with id {parameter_set_id} does not exist'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                serializer = ParameterSetForStorageSerializer(data=param_set_data)
                if serializer.is_valid():
                    parameter_set = serializer.save()
                    parameter_sets_for_storage_ids.append(parameter_set.id)
                else:
                    logger.error("ParameterSetForStorage Serializer Errors: %s", serializer.errors)
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = {
            'room': room_data,
            'responsible': responsible_data,
            'measurement_instruments': measurement_instruments,
            'parameter_sets': parameter_sets_data,
            'extended_parameter_sets': extended_parameter_sets_data,
            'parameter_sets_for_storage': parameter_sets_for_storage_data,
            'created_at': request.data.get('created_at')
        }
        
        serializer = EnvironmentalParametersSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            logger.info("Успешно создана запись с параметрами окружающей среды")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            logger.error("EnvironmentalParametersSerializer Errors: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except Room.DoesNotExist:
        logger.error("Room not found")
        return Response({'error': 'Room not found'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка во время выполнения createEnvironmentalParameters: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def updateEnvironmentalParameters(request, pk):
#     try:
#         environmental_params = EnviromentalParameters.objects.get(pk=pk)
#     except EnviromentalParameters.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = EnvironmentalParametersSerializer(instance=environmental_params, data=request.data, context={'request': request})
    
#     modified_by_data = request.data.get('modified_by')
#     user_id = modified_by_data.get('user')
#     try:
#         modified_by_user = User.objects.get(id=user_id)
#         environmental_params.modified_by = modified_by_user
#     except User.DoesNotExist:
#         print(f'Пользователь с id {user_id} не существует.')
    
#     if serializer.is_valid():
#         room_data = request.data.get('room')
#         room_instance, _ = Room.objects.get_or_create(room_number=room_data.get('room_number'))
#         # Получаем дату из запроса
#         created_at_data = request.data.get('created_at')
#         # Проверяем, существует ли запись для данного помещения на эту дату
#         existing_parameters = EnviromentalParameters.objects.filter(room=room_instance, created_at=created_at_data).exclude(pk=pk)
#         if existing_parameters.exists():
#             print('Вы уже создавали запись на указанную дату')
#             return Response({'error': 'An entry for this room and date already exists'}, status=status.HTTP_400_BAD_REQUEST)

#         has_additional_parameters = room_instance.has_additional_parameters if room_instance else False
#         if has_additional_parameters:
#             measurement_instrument_data = request.data.get('measurement_instrument')
#             measurement_instruments_data = request.data.get('measurement_instruments')
#             created_at_data = request.data.get('created_at')
#             extended_parameter_sets_data = request.data.get('extended_parameter_sets', [])
#             measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)
#             measurement_instruments = []

#             if measurement_instruments_data:
#                 for instrument_data in measurement_instruments_data:
#                     measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
#                     measurement_instruments.append(measurement_instrument_instance)

#             extended_parameter_sets = []

#             for extended_param_set_data in extended_parameter_sets_data:
#                 extended_parameter_set = ExtendedParameterSet.objects.create(
#                     temperature_celsius=extended_param_set_data.get('temperature_celsius'),
#                     humidity_percentage=extended_param_set_data.get('humidity_percentage'),
#                     pressure_kpa=extended_param_set_data.get('pressure_kpa'),
#                     pressure_mmhg=extended_param_set_data.get('pressure_mmhg'),
#                     time=extended_param_set_data.get('time'),
#                     voltage=extended_param_set_data.get('voltage'),
#                     frequency=extended_param_set_data.get('frequency'),
#                     radiation=extended_param_set_data.get('radiation')
#                 )
#                 extended_parameter_sets.append(extended_parameter_set)

#             environmental_params.measurement_instruments.set(measurement_instruments)
#             environmental_params.extended_parameter_sets.set(extended_parameter_sets)
#             environmental_params.created_at = created_at_data
#             environmental_params.save()
#         elif room_instance.is_storage:
#             print('room_instance.is_storage:', room_instance.is_storage)
#             measurement_instrument_data = request.data.get('measurement_instrument')
#             measurement_instruments_data = request.data.get('measurement_instruments')
#             created_at_data = request.data.get('created_at')
#             parameter_sets_for_storage_data = request.data.get('parameter_sets_for_storage', [])
            
#             measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)

#             measurement_instruments = []

#             if measurement_instruments_data:
#                 for instrument_data in measurement_instruments_data:
#                     measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
#                     measurement_instruments.append(measurement_instrument_instance)
#             # Обработка параметрсетов для КВХ
#             parameter_sets_for_storage = []

#             for param_set_data in parameter_sets_for_storage_data:
#                 parameter_set_for_storage = ParameterSetForStorage.objects.create(
#                     temperature_celsius=param_set_data.get('temperature_celsius'),
#                     humidity_percentage=param_set_data.get('humidity_percentage'),
#                     time=param_set_data.get('time')
#                 )
#                 parameter_sets_for_storage.append(parameter_set_for_storage)
            
#             environmental_params.measurement_instruments.set(measurement_instruments)
#             environmental_params.parameter_sets_for_storage.set(parameter_sets_for_storage)
#             environmental_params.created_at = created_at_data
#             environmental_params.save()
#         else:
#             measurement_instrument_data = request.data.get('measurement_instrument')
#             measurement_instruments_data = request.data.get('measurement_instruments')
#             created_at_data = request.data.get('created_at')
#             parameter_sets_data = request.data.get('parameter_sets', [])
            
#             measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)

#             measurement_instruments = []

#             if measurement_instruments_data:
#                 for instrument_data in measurement_instruments_data:
#                     measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
#                     measurement_instruments.append(measurement_instrument_instance)

#             parameter_sets = []

#             for param_set_data in parameter_sets_data:
#                 parameter_set = ParameterSet.objects.create(
#                     temperature_celsius=param_set_data.get('temperature_celsius'),
#                     humidity_percentage=param_set_data.get('humidity_percentage'),
#                     pressure_kpa=param_set_data.get('pressure_kpa'),
#                     pressure_mmhg=param_set_data.get('pressure_mmhg'),
#                     time=param_set_data.get('time')
#                 )
#                 parameter_sets.append(parameter_set)

#             environmental_params.measurement_instruments.set(measurement_instruments)
#             environmental_params.parameter_sets.set(parameter_sets)
#             environmental_params.created_at = created_at_data
#             environmental_params.save()
            
#         return Response(serializer.data, status=status.HTTP_200_OK)
    
#     print("Serializer Errors:", serializer.errors)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateEnvironmentalParameters(request, pk):
    try:
        logger.info(f'Запрос на обновление параметров окружающей среды с ID {pk}')
        environmental_params = EnviromentalParameters.objects.get(pk=pk)
    except EnviromentalParameters.DoesNotExist:
        logger.error(f'Параметры окружающей среды с ID {pk} не найдены')
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = EnvironmentalParametersSerializer(instance=environmental_params, data=request.data, context={'request': request})
    
    modified_by_data = request.data.get('modified_by')
    user_id = modified_by_data.get('user')
    try:
        modified_by_user = User.objects.get(id=user_id)
        environmental_params.modified_by = modified_by_user
    except User.DoesNotExist:
        logger.error(f'Пользователь с ID {user_id} не существует')

    if serializer.is_valid():
        room_data = request.data.get('room')
        room_instance, _ = Room.objects.get_or_create(room_number=room_data.get('room_number'))
        created_at_data = request.data.get('created_at')
        
        existing_parameters = EnviromentalParameters.objects.filter(room=room_instance, created_at=created_at_data).exclude(pk=pk)
        if existing_parameters.exists():
            logger.warning('Запись на указанную дату уже существует')
            return Response({'error': 'An entry for this room and date already exists'}, status=status.HTTP_400_BAD_REQUEST)

        has_additional_parameters = room_instance.has_additional_parameters if room_instance else False
        if has_additional_parameters:
            measurement_instrument_data = request.data.get('measurement_instrument')
            measurement_instruments_data = request.data.get('measurement_instruments')
            created_at_data = request.data.get('created_at')
            extended_parameter_sets_data = request.data.get('extended_parameter_sets', [])
            measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)
            measurement_instruments = []

            if measurement_instruments_data:
                for instrument_data in measurement_instruments_data:
                    measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
                    measurement_instruments.append(measurement_instrument_instance)

            extended_parameter_sets = []

            for extended_param_set_data in extended_parameter_sets_data:
                extended_parameter_set = ExtendedParameterSet.objects.create(
                    temperature_celsius=extended_param_set_data.get('temperature_celsius'),
                    humidity_percentage=extended_param_set_data.get('humidity_percentage'),
                    pressure_kpa=extended_param_set_data.get('pressure_kpa'),
                    pressure_mmhg=extended_param_set_data.get('pressure_mmhg'),
                    time=extended_param_set_data.get('time'),
                    voltage=extended_param_set_data.get('voltage'),
                    frequency=extended_param_set_data.get('frequency'),
                    radiation=extended_param_set_data.get('radiation')
                )
                extended_parameter_sets.append(extended_parameter_set)

            environmental_params.measurement_instruments.set(measurement_instruments)
            environmental_params.extended_parameter_sets.set(extended_parameter_sets)
            environmental_params.created_at = created_at_data
            environmental_params.save()
        elif room_instance.is_storage:
            measurement_instrument_data = request.data.get('measurement_instrument')
            measurement_instruments_data = request.data.get('measurement_instruments')
            created_at_data = request.data.get('created_at')
            parameter_sets_for_storage_data = request.data.get('parameter_sets_for_storage', [])
            
            measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)

            measurement_instruments = []

            if measurement_instruments_data:
                for instrument_data in measurement_instruments_data:
                    measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
                    measurement_instruments.append(measurement_instrument_instance)

            parameter_sets_for_storage = []

            for param_set_data in parameter_sets_for_storage_data:
                parameter_set_for_storage = ParameterSetForStorage.objects.create(
                    temperature_celsius=param_set_data.get('temperature_celsius'),
                    humidity_percentage=param_set_data.get('humidity_percentage'),
                    time=param_set_data.get('time')
                )
                parameter_sets_for_storage.append(parameter_set_for_storage)
            
            environmental_params.measurement_instruments.set(measurement_instruments)
            environmental_params.parameter_sets_for_storage.set(parameter_sets_for_storage)
            environmental_params.created_at = created_at_data
            environmental_params.save()
        else:
            measurement_instrument_data = request.data.get('measurement_instrument')
            measurement_instruments_data = request.data.get('measurement_instruments')
            created_at_data = request.data.get('created_at')
            parameter_sets_data = request.data.get('parameter_sets', [])
            
            measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)

            measurement_instruments = []

            if measurement_instruments_data:
                for instrument_data in measurement_instruments_data:
                    measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
                    measurement_instruments.append(measurement_instrument_instance)

            parameter_sets = []

            for param_set_data in parameter_sets_data:
                parameter_set = ParameterSet.objects.create(
                    temperature_celsius=param_set_data.get('temperature_celsius'),
                    humidity_percentage=param_set_data.get('humidity_percentage'),
                    pressure_kpa=param_set_data.get('pressure_kpa'),
                    pressure_mmhg=param_set_data.get('pressure_mmhg'),
                    time=param_set_data.get('time')
                )
                parameter_sets.append(parameter_set)

            environmental_params.measurement_instruments.set(measurement_instruments)
            environmental_params.parameter_sets.set(parameter_sets)
            environmental_params.created_at = created_at_data
            environmental_params.save()
            
        logger.info(f'Параметры окружающей среды с ID {pk} успешно обновлены')
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    logger.error("Serializer Errors: %s", serializer.errors)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    
# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def deleteEnvironmentalParameters(request, pk):
#     """
#     Удаляет существующий набор параметров окружающей среды.

#     Args:
#         request (Request): Объект HTTP-запроса.
#         pk (int): Первичный ключ параметров окружающей среды.

#     Returns:
#         Response: JSON-ответ, указывающий на успешное или неудачное выполнение операции.
#     """
#     try:
#         environmental_params = EnviromentalParameters.objects.get(pk=pk)
#     except EnviromentalParameters.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     environmental_params.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteEnvironmentalParameters(request, pk):
    """
    Удаляет существующий набор параметров окружающей среды.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ параметров окружающей среды.

    Returns:
        Response: JSON-ответ, указывающий на успешное или неудачное выполнение операции.
    """
    try:
        logger.info(f'Запрос на удаление параметров окружающей среды с ID {pk}')
        environmental_params = EnviromentalParameters.objects.get(pk=pk)
    except EnviromentalParameters.DoesNotExist:
        logger.error(f'Параметры окружающей среды с ID {pk} не найдены')
        return Response(status=status.HTTP_404_NOT_FOUND)

    environmental_params.delete()
    logger.info(f'Параметры окружающей среды с ID {pk} успешно удалены')
    return Response(status=status.HTTP_204_NO_CONTENT)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getParameterSets(request):
#     parameter_sets = ParameterSet.objects.all()
#     serializer = ParameterSetSerializer(parameter_sets, many=True, context={'request': request})
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getParameterSets(request):
    """
    Получает список всех наборов параметров.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ с данными наборов параметров.
    """
    try:
        logger.info('Запрос на получение всех наборов параметров')
        parameter_sets = ParameterSet.objects.all()
        serializer = ParameterSetSerializer(parameter_sets, many=True, context={'request': request})
        logger.info('Наборы параметров успешно получены')
        return Response(serializer.data)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении наборов параметров: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getParameterSet(request, pk):
#     parameter_set = ParameterSet.objects.get(id=pk)
#     serializer = ParameterSetSerializer(parameter_set, many=False)
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getParameterSet(request, pk):
    """
    Получает конкретный набор параметров по первичному ключу.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора параметров.

    Returns:
        Response: JSON-ответ с данными набора параметров.
    """
    try:
        logger.info(f'Запрос на получение набора параметров с id: {pk}')
        parameter_set = ParameterSet.objects.get(id=pk)
        serializer = ParameterSetSerializer(parameter_set, many=False)
        logger.info(f'Набор параметров с id: {pk} успешно получен')
        return Response(serializer.data)
    except ParameterSet.DoesNotExist:
        logger.error(f'Набор параметров с id: {pk} не найден')
        return Response({'error': 'ParameterSet not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении набора параметров с id: {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def createParameterSet(request):
#     # print(request.data)
#     # Преобразовать время в формат 'HH:MM:SS'
#     time_str = request.data.get('time')
#     if time_str:
#         try:
#             datetime.strptime(time_str, '%H:%M:%S')
#         except ValueError:
#             return Response({'error': 'Invalid time format'}, status=status.HTTP_400_BAD_REQUEST)

#     data = request.data
#     if isinstance(data, list):
#         created_sets = []
#         for item in data:
#             serializer = ParameterSetSerializer(data=item)
#             print(serializer.is_valid())
#             if serializer.is_valid():
#                 parameter_set = serializer.save()
#                 created_sets.append(parameter_set)
#             else:
#                 return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#         return Response(ParameterSetSerializer(created_sets, many=True).data, status=status.HTTP_201_CREATED)
#     else:
#         serializer = ParameterSetSerializer(data=data)
#         if serializer.is_valid():
#             parameter_set = serializer.save()
#             return Response(ParameterSetSerializer(parameter_set).data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createParameterSet(request):
    """
    Создает новый набор параметров.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ с данными созданного набора параметров.
    """
    try:
        logger.info('Запрос на создание нового набора параметров')
        time_str = request.data.get('time')
        if time_str:
            try:
                datetime.strptime(time_str, '%H:%M:%S')
            except ValueError:
                logger.error('Неверный формат времени')
                return Response({'error': 'Invalid time format'}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data
        if isinstance(data, list):
            created_sets = []
            for item in data:
                serializer = ParameterSetSerializer(data=item)
                if serializer.is_valid():
                    parameter_set = serializer.save()
                    created_sets.append(parameter_set)
                else:
                    logger.error('Ошибка валидации данных', extra={'errors': serializer.errors})
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            logger.info('Наборы параметров успешно созданы')
            return Response(ParameterSetSerializer(created_sets, many=True).data, status=status.HTTP_201_CREATED)
        else:
            serializer = ParameterSetSerializer(data=data)
            if serializer.is_valid():
                parameter_set = serializer.save()
                logger.info('Набор параметров успешно создан')
                return Response(ParameterSetSerializer(parameter_set).data, status=status.HTTP_201_CREATED)
            logger.error('Ошибка валидации данных', extra={'errors': serializer.errors})
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка при создании набора параметров: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def updateParameterSet(request, pk):
#     try:
#         parameter_set = ParameterSet.objects.get(pk=pk)
#     except ParameterSet.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = ParameterSetSerializer(instance=parameter_set, data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateParameterSet(request, pk):
    """
    Обновляет существующий набор параметров.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора параметров.

    Returns:
        Response: JSON-ответ с обновленными данными набора параметров.
    """
    try:
        logger.info(f'Запрос на обновление набора параметров с ID {pk}')
        try:
            parameter_set = ParameterSet.objects.get(pk=pk)
        except ParameterSet.DoesNotExist:
            logger.error(f'Набор параметров с ID {pk} не найден')
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = ParameterSetSerializer(instance=parameter_set, data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'Набор параметров с ID {pk} успешно обновлен')
            return Response(serializer.data)
        logger.error('Ошибка валидации данных', extra={'errors': serializer.errors})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка при обновлении набора параметров с ID {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def deleteParameterSet(request, pk):
#     try:
#         parameter_set = ParameterSet.objects.get(pk=pk)
#     except ParameterSet.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     parameter_set.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteParameterSet(request, pk):
    """
    Удаляет существующий набор параметров.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора параметров.

    Returns:
        Response: JSON-ответ, указывающий на успешное или неудачное выполнение операции.
    """
    try:
        logger.info(f'Запрос на удаление набора параметров с ID {pk}')
        try:
            parameter_set = ParameterSet.objects.get(pk=pk)
        except ParameterSet.DoesNotExist:
            logger.error(f'Набор параметров с ID {pk} не найден')
            return Response(status=status.HTTP_404_NOT_FOUND)

        parameter_set.delete()
        logger.info(f'Набор параметров с ID {pk} успешно удален')
        return Response(status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        logger.error(f'Произошла ошибка при удалении набора параметров с ID {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getExtendedParameterSets(request):
#     parameter_sets = ExtendedParameterSet.objects.all()
#     serializer = ExtendedParameterSetSerializer(parameter_sets, many=True, context={'request': request})
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getExtendedParameterSets(request):
    """
    Получает список всех наборов расширенных параметров.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий список всех наборов расширенных параметров.
    """
    try:
        logger.info('Запрос на получение всех наборов расширенных параметров')
        parameter_sets = ExtendedParameterSet.objects.all()
        serializer = ExtendedParameterSetSerializer(parameter_sets, many=True, context={'request': request})
        logger.info('Успешно получен список всех наборов расширенных параметров')
        return Response(serializer.data)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении наборов расширенных параметров: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getExtendedParameterSet(request, pk):
#     parameter_set = ExtendedParameterSet.objects.get(id=pk)
#     serializer = ExtendedParameterSetSerializer(parameter_set, many=False)
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getExtendedParameterSet(request, pk):
    """
    Получает конкретный набор расширенных параметров по его первичному ключу.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора расширенных параметров.

    Returns:
        Response: JSON-ответ, содержащий данные набора расширенных параметров.
    """
    try:
        logger.info(f'Запрос на получение набора расширенных параметров с id: {pk}')
        parameter_set = ExtendedParameterSet.objects.get(id=pk)
        serializer = ExtendedParameterSetSerializer(parameter_set, many=False)
        logger.info(f'Успешно получен набор расширенных параметров с id: {pk}')
        return Response(serializer.data)
    except ExtendedParameterSet.DoesNotExist:
        logger.error(f'Набор расширенных параметров с id: {pk} не найден')
        return Response({'error': 'Extended Parameter Set not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении набора расширенных параметров с id: {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def createExtendedParameterSet(request):
#     serializer = ExtendedParameterSetSerializer(data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data, status=status.HTTP_201_CREATED)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createExtendedParameterSet(request):
    """
    Создает новый набор расширенных параметров.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ, содержащий данные созданного набора расширенных параметров.
    """
    try:
        logger.info('Запрос на создание нового набора расширенных параметров')
        serializer = ExtendedParameterSetSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info('Новый набор расширенных параметров успешно создан')
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            logger.error(f'Ошибка валидации данных при создании набора расширенных параметров: {serializer.errors}')
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка при создании набора расширенных параметров: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def updateExtendedParameterSet(request, pk):
#     try:
#         parameter_set = ExtendedParameterSet.objects.get(pk=pk)
#     except ExtendedParameterSet.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = ExtendedParameterSetSerializer(instance=parameter_set, data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateExtendedParameterSet(request, pk):
    """
    Обновляет существующий набор расширенных параметров.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ расширенных параметров.

    Returns:
        Response: JSON-ответ с обновленными данными расширенных параметров.
    """
    try:
        logger.info(f'Запрос на обновление набора расширенных параметров с id: {pk}')
        try:
            parameter_set = ExtendedParameterSet.objects.get(pk=pk)
        except ExtendedParameterSet.DoesNotExist:
            logger.error(f'Набор расширенных параметров с id {pk} не найден')
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = ExtendedParameterSetSerializer(instance=parameter_set, data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'Набор расширенных параметров с id {pk} успешно обновлен')
            return Response(serializer.data)
        else:
            logger.error(f'Ошибка валидации данных при обновлении набора расширенных параметров с id {pk}: {serializer.errors}')
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка при обновлении набора расширенных параметров с id {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def deleteExtendedParameterSet(request, pk):
#     try:
#         parameter_set = ExtendedParameterSet.objects.get(pk=pk)
#     except ExtendedParameterSet.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     parameter_set.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteExtendedParameterSet(request, pk):
    """
    Удаляет существующий набор расширенных параметров.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ расширенных параметров.

    Returns:
        Response: JSON-ответ, указывающий на успешное или неудачное выполнение операции.
    """
    try:
        logger.info(f'Запрос на удаление набора расширенных параметров с id: {pk}')
        try:
            parameter_set = ExtendedParameterSet.objects.get(pk=pk)
        except ExtendedParameterSet.DoesNotExist:
            logger.error(f'Набор расширенных параметров с id {pk} не найден')
            return Response(status=status.HTTP_404_NOT_FOUND)

        parameter_set.delete()
        logger.info(f'Набор расширенных параметров с id {pk} успешно удален')
        return Response(status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        logger.error(f'Произошла ошибка при удалении набора расширенных параметров с id {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getStorageParameterSets(request):
#     parameter_sets = ParameterSetForStorage.objects.all()
#     serializer = ParameterSetForStorageSerializer(parameter_sets, many=True, context={'request': request})
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getStorageParameterSets(request):
    """
    Получает все наборы параметров для КВХ.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ с наборами параметров для КВХ.
    """
    try:
        logger.info('Запрос на получение всех наборов параметров для КВХ')
        parameter_sets = ParameterSetForStorage.objects.all()
        serializer = ParameterSetForStorageSerializer(parameter_sets, many=True, context={'request': request})
        logger.info('Наборы параметров для КВХ успешно получены')
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении наборов параметров для КВХ: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getStorageParameterSet(request, pk):
#     parameter_set = ParameterSetForStorage.objects.get(id=pk)
#     serializer = ParameterSetForStorageSerializer(parameter_set, many=False)
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getStorageParameterSet(request, pk):
    """
    Получает конкретный набор параметров для КВХ по идентификатору.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора параметров для КВХ.

    Returns:
        Response: JSON-ответ с набором параметров для КВХ.
    """
    try:
        logger.info(f'Запрос на получение набора параметров для КВХ с id={pk}')
        parameter_set = ParameterSetForStorage.objects.get(id=pk)
        serializer = ParameterSetForStorageSerializer(parameter_set, many=False)
        logger.info(f'Набор параметров для КВХ с id={pk} успешно получен')
        return Response(serializer.data)
    except ParameterSetForStorage.DoesNotExist:
        logger.error(f'Набор параметров для КВХ с id={pk} не найден')
        return Response({'error': 'Набор параметров для КВХ не найден'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f'Произошла ошибка при получении набора параметров для КВХ с id={pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def createStorageParameterSet(request):
#     serializer = ParameterSetForStorageSerializer(data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data, status=status.HTTP_201_CREATED)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createStorageParameterSet(request):
    """
    Создает новый набор параметров для КВХ.

    Args:
        request (Request): Объект HTTP-запроса.

    Returns:
        Response: JSON-ответ с созданным набором параметров для КВХ или ошибками валидации.
    """
    try:
        logger.info('Запрос на создание нового набора параметров для КВХ')
        serializer = ParameterSetForStorageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info('Набор параметров для КВХ успешно создан')
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.warning(f'Ошибки валидации при создании набора параметров для КВХ: {serializer.errors}')
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка при создании набора параметров для КВХ: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def updateStorageParameterSet(request, pk):
#     try:
#         parameter_set = ParameterSetForStorage.objects.get(pk=pk)
#     except ParameterSetForStorage.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = ParameterSetForStorageSerializer(instance=parameter_set, data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateStorageParameterSet(request, pk):
    """
    Обновляет существующий набор параметров для КВХ.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора параметров для КВХ.

    Returns:
        Response: JSON-ответ с обновленным набором параметров для хранения или ошибками валидации.
    """
    try:
        logger.info(f'Запрос на обновление набора параметров для КВХ с ID {pk}')
        try:
            parameter_set = ParameterSetForStorage.objects.get(pk=pk)
        except ParameterSetForStorage.DoesNotExist:
            logger.warning(f'Набор параметров для КВХ с ID {pk} не найден')
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = ParameterSetForStorageSerializer(instance=parameter_set, data=request.data)
        if serializer.is_valid():
            serializer.save()
            logger.info(f'Набор параметров для КВХ с ID {pk} успешно обновлен')
            return Response(serializer.data)
        logger.warning(f'Ошибки валидации при обновлении набора параметров для КВХ с ID {pk}: {serializer.errors}')
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Произошла ошибка при обновлении набора параметров для КВХ с ID {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def deleteStorageParameterSet(request, pk):
#     try:
#         parameter_set = ParameterSetForStorage.objects.get(pk=pk)
#     except ParameterSetForStorage.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     parameter_set.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteStorageParameterSet(request, pk):
    """
    Удаляет существующий набор параметров для КВХ.

    Args:
        request (Request): Объект HTTP-запроса.
        pk (int): Первичный ключ набора параметров для КВХ.

    Returns:
        Response: JSON-ответ, указывающий на успешное или неудачное выполнение операции.
    """
    try:
        logger.info(f'Запрос на удаление набора параметров для КВХ с ID {pk}')
        try:
            parameter_set = ParameterSetForStorage.objects.get(pk=pk)
        except ParameterSetForStorage.DoesNotExist:
            logger.warning(f'Набор параметров для КВХ с ID {pk} не найден')
            return Response(status=status.HTTP_404_NOT_FOUND)

        parameter_set.delete()
        logger.info(f'Набор параметров для КВХ с ID {pk} успешно удален')
        return Response(status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        logger.error(f'Произошла ошибка при удалении набора параметров для КВХ с ID {pk}: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Выгрузка в Excel данных о параметрах для помещений

@api_view(['GET'])
def export_parameters_to_excel(request):
    try:
        logger.info(f"Export request from user: {request.user.username}")

        # Создание книги Excel
        wb = Workbook()
        ws = wb.active

        # Добавление заголовков
        columns = ['Помещение', 'Ответственный', 'Средства измерений', 'Время создания записи', 'Кто создал', 'Время изменения', 'Кто изменил',
                   'Температура (°C)', 'Влажность (%)', 'Давление (кПа)', 'Давление (мм рт. ст.)', 'Напряжение (В)', 'Частота (Гц)', 'Радиационный фон', 'Время создания набора параметров']
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        row_num = 2

        for param in EnviromentalParameters.objects.all():
            # Добавляем данные о параметрах
            ws.append([
                param.room.room_number,
                f'{param.responsible.last_name} {param.responsible.first_name} {param.responsible.patronymic}' if param.responsible else '',
                ', '.join([f'{instrument.name} {instrument.type} ({instrument.serial_number})' for instrument in param.measurement_instruments.all()]) if param.measurement_instruments.exists() else '',
                param.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                param.created_by.username if param.created_by else '',
                param.modified_at.strftime('%Y-%m-%d %H:%M:%S'),
                param.modified_by.username if param.modified_by else '',
                '', '', '', '', '', '', '', '', '',  # Добавляем пустые значения для температуры, влажности, давления и т.д.
            ])
            row_num += 1

            # Добавляем параметры из parameter_sets
            for param_set in param.parameter_sets.all():
                ws.append([
                    '', '', '', '', '', '', '',  # Добавляем пустые значения для комнаты, ответственного, инструмента и т.д.
                    param_set.temperature_celsius if param_set.temperature_celsius is not None else '',
                    param_set.humidity_percentage if param_set.humidity_percentage is not None else '',
                    param_set.pressure_kpa if param_set.pressure_kpa is not None else '',
                    param_set.pressure_mmhg if param_set.pressure_mmhg is not None else '',
                    '', '', '',  # Добавляем пустые значения для Voltage (V), Frequency (Hz), Radiation
                    param_set.time.strftime('%H:%M:%S') if param_set.time else '',
                ])
                row_num += 1

            # Добавляем параметры из extended_parameter_sets
            for extended_param_set in param.extended_parameter_sets.all():
                ws.append([
                    '', '', '', '', '', '', '',  # Добавляем пустые значения для комнаты, ответственного, инструмента и т.д.
                    extended_param_set.temperature_celsius if extended_param_set.temperature_celsius is not None else '',
                    extended_param_set.humidity_percentage if extended_param_set.humidity_percentage is not None else '',
                    extended_param_set.pressure_kpa if extended_param_set.pressure_kpa is not None else '',
                    extended_param_set.pressure_mmhg if extended_param_set.pressure_mmhg is not None else '',
                    extended_param_set.voltage if extended_param_set.voltage is not None else '',  # Отображаем напряжение
                    extended_param_set.frequency if extended_param_set.frequency is not None else '',  # Отображаем частоту
                    extended_param_set.radiation if extended_param_set.radiation is not None else '', 
                    extended_param_set.time.strftime('%H:%M:%S') if extended_param_set.time else '',
                ])
                row_num += 1
                
            # Добавляем параметры из parameter_sets_for_storage
            for param_set_for_storage in param.parameter_sets_for_storage.all():
                ws.append([
                    '', '', '', '', '', '', '',  # Добавляем пустые значения для комнаты, ответственного, инструмента и т.д.
                    param_set_for_storage.temperature_celsius if param_set_for_storage.temperature_celsius is not None else '',
                    param_set_for_storage.humidity_percentage if param_set_for_storage.humidity_percentage is not None else '',
                    '', '', '', '', '',  # Добавляем пустые значения для pressure_kpa(kPa), pressure_mmhg(mmHg), Voltage (V), Frequency (Hz), Radiation
                    param_set_for_storage.time.strftime('%H:%M:%S') if param_set_for_storage.time else '',
                ])
                row_num += 1

            row_num += 1  # Переходим на следующую строку для следующего параметра

        # Создание HTTP-ответа
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="environmental_parameters.xlsx"'

        # Использование BytesIO для сохранения в памяти
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response.write(buffer.getvalue())
        buffer.close()

        return response

    except Exception as e:
        logger.error(f"Error exporting parameters: {str(e)}")
        return HttpResponseServerError("Internal Server Error")

# Выгрузка в Excel данных о параметрах для зданий

@api_view(['GET'])
def export_parameters_for_buildings_to_excel(request):
    try:
        logger.info(f"Export request from user: {request.user.username}")

        # Создание книги Excel
        wb = Workbook()
        ws = wb.active

        # Добавление заголовков
        columns = ['Здание', 'Ответственный', 'Средства измерений', 'Время создания записи', 'Кто создал', 'Время изменения', 'Кто изменил',
                   'Напряжение (В)', 'Частота (Гц)', 'Время создания набора параметров']
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        row_num = 2

        for param in BuildingEnviromentalParameters.objects.all():
            # Добавляем данные о параметрах
            ws.append([
                param.building.building_number,
                f'{param.responsible.last_name} {param.responsible.first_name} {param.responsible.patronymic}' if param.responsible else '',
                ', '.join([f'{instrument.name} {instrument.type} ({instrument.serial_number})' for instrument in param.measurement_instruments.all()]) if param.measurement_instruments.exists() else '',
                param.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                param.created_by.username if param.created_by else '',
                param.modified_at.strftime('%Y-%m-%d %H:%M:%S'),
                param.modified_by.username if param.modified_by else '',
                '', '', '', '',  # Добавляем пустые значения для напряжения, частоты и т.д.
            ])
            row_num += 1

            # Добавляем параметры из parameter_sets
            for param_set in param.parameter_sets.all():
                ws.append([
                    '', '', '', '', '', '', '',  # Добавляем пустые значения для комнаты, ответственного, инструмента и т.д.
                    param_set.voltage if param_set.voltage is not None else '',  # Отображаем напряжение
                    param_set.frequency if param_set.frequency is not None else '',  # Отображаем частоту
                    param_set.time.strftime('%H:%M:%S') if param_set.time else '',
                ])
                row_num += 1

            row_num += 1  # Переходим на следующую строку для следующего параметра

        # Создание HTTP-ответа
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="building_environmental_parameters.xlsx"'

        # Использование BytesIO для сохранения в памяти
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response.write(buffer.getvalue())
        buffer.close()

        return response

    except Exception as e:
        logger.error(f"Error exporting parameters: {str(e)}")
        return HttpResponseServerError("Internal Server Error")

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getBuildingParameterSets(request):
#     parameter_sets = BuildingParameterSet.objects.all()
#     serializer = BuildingParameterSetSerializer(parameter_sets, many=True, context={'request': request})
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getBuildingParameterSets(request):
    try:
        logger.info(f"Пользователь {request.user.username} запросил все наборы параметров зданий")
        parameter_sets = BuildingParameterSet.objects.all()
        serializer = BuildingParameterSetSerializer(parameter_sets, many=True, context={'request': request})
        logger.info(f"Успешно получены и сериализованы {len(parameter_sets)} наборы параметров зданий")
        return Response(serializer.data)
    except Exception as e:
        logger.error(f"Ошибка при получении наборов параметров зданий: {str(e)}")
        return Response({"detail": "Внутренняя ошибка сервера"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getBuildingParameterSet(request, pk):
#     parameter_set = BuildingParameterSet.objects.get(id=pk)
#     serializer = BuildingParameterSetSerializer(parameter_set, many=False)
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getBuildingParameterSet(request, pk):
    try:
        logger.info(f"Пользователь {request.user.username} запросил набор параметров здания с id: {pk}")
        parameter_set = BuildingParameterSet.objects.get(id=pk)
        serializer = BuildingParameterSetSerializer(parameter_set, many=False)
        logger.info(f"Успешно получен и сериализован набор параметров здания с id: {pk}")
        return Response(serializer.data)
    except BuildingParameterSet.DoesNotExist:
        logger.warning(f"Набор параметров здания с id: {pk} не найден")
        return Response({"detail": "Набор параметров не найден"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Ошибка при получении набора параметров здания с id: {pk}: {str(e)}")
        return Response({"detail": "Внутренняя ошибка сервера"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def createBuildingParameterSet(request):
#     print(request.data)
#     # Преобразовать время в формат 'HH:MM:SS'
#     time_str = request.data.get('time')
#     if time_str:
#         try:
#             datetime.strptime(time_str, '%H:%M:%S')
#         except ValueError:
#             return Response({'error': 'Invalid time format'}, status=status.HTTP_400_BAD_REQUEST)

#     data = request.data
#     if isinstance(data, list):
#         created_sets = []
#         for item in data:
#             serializer = BuildingParameterSetSerializer(data=item)
#             print(serializer.is_valid())
#             if serializer.is_valid():
#                 parameter_set = serializer.save()
#                 created_sets.append(parameter_set)
#             else:
#                 return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#         return Response(BuildingParameterSetSerializer(created_sets, many=True).data, status=status.HTTP_201_CREATED)
#     else:
#         serializer = BuildingParameterSetSerializer(data=data)
#         if serializer.is_valid():
#             parameter_set = serializer.save()
#             return Response(BuildingParameterSetSerializer(parameter_set).data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createBuildingParameterSet(request):
    try:
        logger.info(f"Пользователь {request.user.username} пытается создать новый набор параметров здания")
        logger.debug(f"Данные запроса: {request.data}")

        # Преобразовать время в формат 'HH:MM:SS'
        time_str = request.data.get('time')
        if time_str:
            try:
                datetime.strptime(time_str, '%H:%M:%S')
            except ValueError:
                logger.warning(f"Некорректный формат времени: {time_str}")
                return Response({'error': 'Invalid time format'}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data
        if isinstance(data, list):
            created_sets = []
            for item in data:
                serializer = BuildingParameterSetSerializer(data=item)
                if serializer.is_valid():
                    parameter_set = serializer.save()
                    created_sets.append(parameter_set)
                    logger.info(f"Успешно создан набор параметров здания с id: {parameter_set.id}")
                else:
                    logger.warning(f"Ошибка валидации данных: {serializer.errors}")
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            return Response(BuildingParameterSetSerializer(created_sets, many=True).data, status=status.HTTP_201_CREATED)
        else:
            serializer = BuildingParameterSetSerializer(data=data)
            if serializer.is_valid():
                parameter_set = serializer.save()
                logger.info(f"Успешно создан набор параметров здания с id: {parameter_set.id}")
                return Response(BuildingParameterSetSerializer(parameter_set).data, status=status.HTTP_201_CREATED)
            logger.warning(f"Ошибка валидации данных: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f"Ошибка при создании набора параметров здания: {str(e)}")
        return Response({'detail': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def updateBuildingParameterSet(request, pk):
#     try:
#         parameter_set = BuildingParameterSet.objects.get(pk=pk)
#     except BuildingParameterSet.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = BuildingParameterSetSerializer(instance=parameter_set, data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(serializer.data)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateBuildingParameterSet(request, pk):
    try:
        logger.info(f"Пользователь {request.user.username} пытается обновить набор параметров здания с id: {pk}")
        parameter_set = BuildingParameterSet.objects.get(pk=pk)
    except BuildingParameterSet.DoesNotExist:
        logger.warning(f"Набор параметров здания с id {pk} не найден")
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = BuildingParameterSetSerializer(instance=parameter_set, data=request.data)
    if serializer.is_valid():
        serializer.save()
        logger.info(f"Набор параметров здания с id {pk} успешно обновлен")
        return Response(serializer.data)
    
    logger.warning(f"Ошибка валидации данных при обновлении набора параметров здания с id {pk}: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def deleteBuildingParameterSet(request, pk):
#     try:
#         parameter_set = BuildingParameterSet.objects.get(pk=pk)
#     except BuildingParameterSet.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     parameter_set.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteBuildingParameterSet(request, pk):
    try:
        logger.info(f"Пользователь {request.user.username} пытается удалить набор параметров здания с id: {pk}")
        parameter_set = BuildingParameterSet.objects.get(pk=pk)
    except BuildingParameterSet.DoesNotExist:
        logger.warning(f"Набор параметров здания с id {pk} не найден")
        return Response(status=status.HTTP_404_NOT_FOUND)

    parameter_set.delete()
    logger.info(f"Набор параметров здания с id {pk} успешно удален")
    return Response(status=status.HTTP_204_NO_CONTENT)


# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def getBuildingEnvironmentalParameters(request):
#     try:
#         user = request.user
#         responsible = request.query_params.get('responsible')
#         building = request.query_params.get('building')
#         date = request.query_params.get('date')

#         parameters = BuildingEnviromentalParameters.objects.all().prefetch_related('building', 'responsible')

#         if responsible:
#             parameters = parameters.filter(responsible=responsible)
#         if building:
#             parameters = parameters.filter(building=building)
#         if date:
#             created_start = datetime.strptime(date, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
#             created_end = created_start + timedelta(days=1)
#             parameters = parameters.filter(created_at__range=(created_start, created_end))

#         parameters = parameters.order_by('-created_at')  # Добавляем сортировку по дате создания записи

#         serializer = BuildingEnvironmentalParametersSerializer(parameters, many=True, context={'request': request})
#         return Response(serializer.data)

#     except Exception as e:
#         logger.error(f'Произошла ошибка во время выполнения getEnviromentalParameters: {e}', exc_info=True)
#         return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def getBuildingEnvironmentalParameters(request):
    try:
        user = request.user
        logger.info(f"Запрос от пользователя: {user.username}")

        responsible = request.query_params.get('responsible')
        building = request.query_params.get('building')
        date = request.query_params.get('date')

        logger.info(f"Фильтры запроса - responsible: {responsible}, building: {building}, date: {date}")

        parameters = BuildingEnviromentalParameters.objects.all().prefetch_related('building', 'responsible')

        if responsible:
            parameters = parameters.filter(responsible=responsible)
            logger.info(f"Фильтрация по ответственному: {responsible}")
        if building:
            parameters = parameters.filter(building=building)
            logger.info(f"Фильтрация по зданию: {building}")
        if date:
            created_start = datetime.strptime(date, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            created_end = created_start + timedelta(days=1)
            parameters = parameters.filter(created_at__range=(created_start, created_end))
            logger.info(f"Фильтрация по дате: {created_start} - {created_end}")

        parameters = parameters.order_by('-created_at')  # Добавляем сортировку по дате создания записи
        logger.info("Параметры отсортированы по дате создания")

        serializer = BuildingEnvironmentalParametersSerializer(parameters, many=True, context={'request': request})
        return Response(serializer.data)

    except Exception as e:
        logger.error(f'Произошла ошибка во время выполнения getBuildingEnvironmentalParameters: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def getBuildingEnvironmentalParameter(request, pk):
#     parameters = BuildingEnviromentalParameters.objects.get(id=pk)
#     serializer = BuildingEnvironmentalParametersSerializer(parameters, many=False)
#     return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getBuildingEnvironmentalParameter(request, pk):
    try:
        user = request.user
        logger.info(f"Запрос от пользователя: {user.username} для получения параметра окружающей среды здания с ID: {pk}")

        parameters = BuildingEnviromentalParameters.objects.get(id=pk)
        logger.info(f"Найден параметр окружающей среды здания с ID: {pk}")

        serializer = BuildingEnvironmentalParametersSerializer(parameters, many=False)
        return Response(serializer.data)
    
    except BuildingEnviromentalParameters.DoesNotExist:
        logger.warning(f"Параметр окружающей среды здания с ID: {pk} не найден")
        return Response({'error': 'Параметр окружающей среды не найден'}, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        logger.error(f'Произошла ошибка во время выполнения getBuildingEnvironmentalParameter: {e}', exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def createBuildingEnvironmentalParameters(request):
#     try:
#         # print("Request Data:", request.data)
#         building_data = request.data.get('building')
#         # print("Building Data:", building_data)
#         building = Building.objects.get(building_number=building_data.get('building_number'))
#         # print("Building:", building)
#         created_at = request.data.get('created_at')
#         existing_parameters = BuildingEnviromentalParameters.objects.filter(building=building, created_at=created_at)
#         if existing_parameters.exists():
#             print('Вы уже создавали запись на указанную дату')
#             return Response({'error': 'An entry for this room and date already exists'}, status=status.HTTP_400_BAD_REQUEST)
#         responsible_data = request.data.get('responsible')
#         # print("Responsible Data:", responsible_data)
#         responsible, _ = Responsible.objects.get_or_create(
#             first_name=responsible_data.get('first_name'),
#             last_name=responsible_data.get('last_name'),
#             patronymic=responsible_data.get('patronymic')
#         )
#         # print("Responsible:", responsible)
#         measurement_instruments_data = request.data.get('measurement_instruments')
#         # print("Measurement Instruments Data:", measurement_instruments_data)
#         measurement_instruments_data = request.data.get('measurement_instruments', [])

#         measurement_instruments = []
#         for instrument_data in measurement_instruments_data:
#             instrument_dict = {
#                 'name': instrument_data.get('name'),  
#                 'type': instrument_data.get('type'),
#                 'serial_number': instrument_data.get('serial_number'),
#                 'calibration_date': instrument_data.get('calibration_date'),
#                 'calibration_interval': instrument_data.get('calibration_interval')
#             }
#             measurement_instruments.append(instrument_dict)

#         # print("Measurement Instruments:", measurement_instruments)
#         parameter_sets_data = request.data.get('parameter_sets', [])
#         # print("Parameter Sets Data:", parameter_sets_data)
#         parameter_set_ids = []
#         building_data = request.data.get('building')
#         # print("Building data:", building_data)
#         building = Building.objects.get(building_number=building_data.get('building_number'))
#         # print("Building:", building)
#         # Создать список id существующих параметров, чтобы избежать дублирования
#         existing_parameter_set_ids = BuildingParameterSet.objects.values_list('id', flat=True)
#         created_at = request.data.get('created_at')
#         existing_parameters = BuildingEnviromentalParameters.objects.filter(building=building, created_at=created_at)
#         if existing_parameters.exists():
#             print('Вы уже создавали запись на указанную дату')
#             return Response({'error': 'An entry for this room and date already exists'}, status=status.HTTP_400_BAD_REQUEST)
#         for param_set_data in parameter_sets_data:
#             parameter_set_id = param_set_data.get('id')
#             # print(f"Получен parameter_set_id: {parameter_set_id}")
#             # Проверить, существует ли параметр с таким id
#             if parameter_set_id in existing_parameter_set_ids:
#                 parameter_set_ids.append(parameter_set_id)
#                 # print(f"Найден ParameterSet с id {parameter_set_id}")
#             else:
#                 serializer = BuildingParameterSetSerializer(data=param_set_data)
#                 if serializer.is_valid():
#                     parameter_set = serializer.save()
#                     parameter_set_ids.append(parameter_set.id)
#                 else:
#                     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
#         data = {
#             'building': building_data,
#             'responsible': responsible_data,
#             'measurement_instruments': measurement_instruments,
#             'parameter_sets': parameter_sets_data,
#             'created_at': request.data.get('created_at')
#         }
#         # print("Data:", data)
#         serializer = BuildingEnvironmentalParametersSerializer(data=data, context={'request': request})
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         else:
#             print("Serializer Errors:", serializer.errors)
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#     except Room.DoesNotExist:
#         print("Room not found")
#         return Response({'error': 'Room not found'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createBuildingEnvironmentalParameters(request):
    try:
        user = request.user
        logger.info(f"Запрос на создание параметров окружающей среды здания от пользователя: {user.username}")

        building_data = request.data.get('building')
        building = Building.objects.get(building_number=building_data.get('building_number'))
        created_at = request.data.get('created_at')
        existing_parameters = BuildingEnviromentalParameters.objects.filter(building=building, created_at=created_at)
        
        if existing_parameters.exists():
            logger.warning('Запись для этого здания и даты уже существует')
            return Response({'error': 'An entry for this building and date already exists'}, status=status.HTTP_400_BAD_REQUEST)

        responsible_data = request.data.get('responsible')
        responsible, _ = Responsible.objects.get_or_create(
            first_name=responsible_data.get('first_name'),
            last_name=responsible_data.get('last_name'),
            patronymic=responsible_data.get('patronymic')
        )

        measurement_instruments_data = request.data.get('measurement_instruments', [])
        measurement_instruments = []
        for instrument_data in measurement_instruments_data:
            instrument_dict = {
                'name': instrument_data.get('name'),  
                'type': instrument_data.get('type'),
                'serial_number': instrument_data.get('serial_number'),
                'calibration_date': instrument_data.get('calibration_date'),
                'calibration_interval': instrument_data.get('calibration_interval')
            }
            measurement_instruments.append(instrument_dict)

        parameter_sets_data = request.data.get('parameter_sets', [])
        parameter_set_ids = []

        existing_parameter_set_ids = BuildingParameterSet.objects.values_list('id', flat=True)
        for param_set_data in parameter_sets_data:
            parameter_set_id = param_set_data.get('id')
            if parameter_set_id in existing_parameter_set_ids:
                parameter_set_ids.append(parameter_set_id)
                logger.info(f"Найден ParameterSet с id {parameter_set_id}")
            else:
                serializer = BuildingParameterSetSerializer(data=param_set_data)
                if serializer.is_valid():
                    parameter_set = serializer.save()
                    parameter_set_ids.append(parameter_set.id)
                    logger.info(f"Создан новый ParameterSet с id {parameter_set.id}")
                else:
                    logger.error(f"Ошибки в данных ParameterSet: {serializer.errors}")
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        data = {
            'building': building_data,
            'responsible': responsible_data,
            'measurement_instruments': measurement_instruments,
            'parameter_sets': parameter_sets_data,
            'created_at': request.data.get('created_at')
        }

        serializer = BuildingEnvironmentalParametersSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            logger.info("Параметры окружающей среды здания успешно созданы")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            logger.error(f"Ошибки в данных BuildingEnvironmentalParameters: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except Building.DoesNotExist:
        logger.error("Здание не найдено")
        return Response({'error': 'Building not found'}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        logger.error(f"Произошла ошибка во время выполнения createBuildingEnvironmentalParameters: {e}", exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def updateBuildingEnvironmentalParameters(request, pk):   
#     try:
#         environmental_params = BuildingEnviromentalParameters.objects.get(pk=pk)
#     except BuildingEnviromentalParameters.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     serializer = BuildingEnvironmentalParametersSerializer(instance=environmental_params, data=request.data, context={'request': request})
#     environmental_params.parameter_sets.all().delete()
#     if serializer.is_valid():
#         building_data = request.data.get('building')
#         parameter_sets_data = request.data.get('parameter_sets', [])
#         measurement_instrument_data = request.data.get('measurement_instrument')
#         measurement_instruments_data = request.data.get('measurement_instruments', [])  # Получаем список средств измерения
#         modified_by_data = request.data.get('modified_by')
#         user_id = modified_by_data.get('user')

#         building, created = Building.objects.get_or_create(building_number=building_data.get('building_number')) if building_data else (None, False)

#         parameter_sets = []
#         for param_set_data in parameter_sets_data:
#             parameter_set = BuildingParameterSet.objects.create(**param_set_data)
#             parameter_sets.append(parameter_set)

#         measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)
#         measurement_instruments = []
        
#         if measurement_instruments_data:
#             for instrument_data in measurement_instruments_data:
#                 measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
#                 measurement_instruments.append(measurement_instrument_instance)

#         environmental_params.building = building
#         created_at = request.data.get('created_at')
#         if created_at:
#             environmental_params.created_at = created_at

#         environmental_params.parameter_sets.set(parameter_sets)
#         environmental_params.measurement_instruments.set(measurement_instruments)

#         try:
#             modified_by_user = User.objects.get(id=user_id)
#             environmental_params.modified_by = modified_by_user
#         except User.DoesNotExist:
#             print(f'Пользователь с id {user_id} не существует.')

#         environmental_params.save()

#         return Response(serializer.data)

#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateBuildingEnvironmentalParameters(request, pk):
    try:
        user = request.user
        logger.info(f"Запрос на обновление параметров окружающей среды здания от пользователя: {user.username}, id параметров: {pk}")

        try:
            environmental_params = BuildingEnviromentalParameters.objects.get(pk=pk)
        except BuildingEnviromentalParameters.DoesNotExist:
            logger.warning(f"Параметры окружающей среды здания с id {pk} не найдены")
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = BuildingEnvironmentalParametersSerializer(instance=environmental_params, data=request.data, context={'request': request})
        environmental_params.parameter_sets.all().delete()
        
        if serializer.is_valid():
            building_data = request.data.get('building')
            parameter_sets_data = request.data.get('parameter_sets', [])
            measurement_instrument_data = request.data.get('measurement_instrument')
            measurement_instruments_data = request.data.get('measurement_instruments', [])
            modified_by_data = request.data.get('modified_by')
            user_id = modified_by_data.get('user')

            building, created = Building.objects.get_or_create(building_number=building_data.get('building_number')) if building_data else (None, False)

            parameter_sets = []
            for param_set_data in parameter_sets_data:
                parameter_set = BuildingParameterSet.objects.create(**param_set_data)
                parameter_sets.append(parameter_set)
                logger.info(f"Создан новый параметр набора с id {parameter_set.id}")

            measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**measurement_instrument_data) if measurement_instrument_data else (None, False)
            measurement_instruments = []
            if measurement_instruments_data:
                for instrument_data in measurement_instruments_data:
                    measurement_instrument_instance, _ = MeasurementInstrument.objects.get_or_create(**instrument_data)
                    measurement_instruments.append(measurement_instrument_instance)

            environmental_params.building = building
            created_at = request.data.get('created_at')
            if created_at:
                environmental_params.created_at = created_at

            environmental_params.parameter_sets.set(parameter_sets)
            environmental_params.measurement_instruments.set(measurement_instruments)

            try:
                modified_by_user = User.objects.get(id=user_id)
                environmental_params.modified_by = modified_by_user
                logger.info(f"Параметры изменены пользователем: {modified_by_user.username}")
            except User.DoesNotExist:
                logger.warning(f'Пользователь с id {user_id} не существует.')

            environmental_params.save()
            logger.info(f"Параметры окружающей среды здания с id {pk} успешно обновлены")
            return Response(serializer.data)

        logger.error(f"Ошибки в данных обновления параметров окружающей среды: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        logger.error(f"Произошла ошибка во время обновления параметров окружающей среды здания: {e}", exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
# def deleteBuildingEnvironmentalParameters(request, pk):
#     try:
#         environmental_params = BuildingEnviromentalParameters.objects.get(pk=pk)
#     except BuildingEnviromentalParameters.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)

#     environmental_params.delete()
#     return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deleteBuildingEnvironmentalParameters(request, pk):
    try:
        user = request.user
        logger.info(f"Запрос на удаление параметров окружающей среды здания от пользователя: {user.username}, id параметров: {pk}")

        try:
            environmental_params = BuildingEnviromentalParameters.objects.get(pk=pk)
        except BuildingEnviromentalParameters.DoesNotExist:
            logger.warning(f"Параметры окружающей среды здания с id {pk} не найдены")
            return Response(status=status.HTTP_404_NOT_FOUND)

        environmental_params.delete()
        logger.info(f"Параметры окружающей среды здания с id {pk} успешно удалены")
        return Response(status=status.HTTP_204_NO_CONTENT)

    except Exception as e:
        logger.error(f"Произошла ошибка во время удаления параметров окружающей среды здания: {e}", exc_info=True)
        return Response({'error': 'Внутренняя ошибка сервера'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# # Фильтрация по помещениям
# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def filterEnvironmentalParameters(request):
#     try:
#         responsible_id = request.query_params.get('responsible')
#         room_number = request.query_params.get('room')
#         date = request.query_params.get('date')
#         start_date = request.query_params.get('start_date')
#         end_date = request.query_params.get('end_date')

#         print("Room number from request:", room_number)  

#         parameters = EnviromentalParameters.objects.all()

#         if responsible_id:
#             parameters = parameters.filter(responsible__id=responsible_id)

#         if room_number:
#             parameters = parameters.filter(room__room_number=room_number)

#         if date:
#             parameters = parameters.filter(created_at=date)

#         if start_date and end_date:
#             parameters = parameters.filter(created_at__range=[start_date, end_date])
        
#         parameters = parameters.order_by('-created_at')

#         serializer = EnvironmentalParametersSerializer(parameters, many=True)
#         return Response(serializer.data)

#     except Exception as e:
#         print("Error filtering parameters:", str(e))
#         return Response({'error': 'An error occurred while filtering parameters'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def filterEnvironmentalParameters(request):
    try:
        user = request.user
        logger.info(f"Запрос на фильтрацию параметров окружающей среды от пользователя: {user.username}")

        responsible_id = request.query_params.get('responsible')
        room_number = request.query_params.get('room')
        date = request.query_params.get('date')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        logger.info(f"Параметры фильтрации: responsible_id={responsible_id}, room_number={room_number}, date={date}, start_date={start_date}, end_date={end_date}")

        parameters = EnviromentalParameters.objects.all()

        if responsible_id:
            parameters = parameters.filter(responsible__id=responsible_id)

        if room_number:
            parameters = parameters.filter(room__room_number=room_number)

        if date:
            parameters = parameters.filter(created_at=date)

        if start_date and end_date:
            parameters = parameters.filter(created_at__range=[start_date, end_date])

        parameters = parameters.order_by('-created_at')

        serializer = EnvironmentalParametersSerializer(parameters, many=True)
        logger.info(f"Фильтрация завершена успешно, найдено {parameters.count()} параметров")
        return Response(serializer.data)

    except Exception as e:
        logger.error(f"Произошла ошибка при фильтрации параметров: {e}", exc_info=True)
        return Response({'error': 'Произошла ошибка при фильтрации параметров'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
# # Фильтрация по зданиям
# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def filterBuildingEnvironmentalParameters(request):
#     try:
#         responsible_id = request.query_params.get('responsible')
#         building_number = request.query_params.get('building')
#         date = request.query_params.get('date')
#         start_date = request.query_params.get('start_date')
#         end_date = request.query_params.get('end_date')
#         print("responsible_id from request:", responsible_id) 

#         print("Building number from request:", building_number)  

#         parameters = BuildingEnviromentalParameters.objects.all()

#         if responsible_id:
#             parameters = parameters.filter(responsible__id=responsible_id)

#         if building_number:
#             parameters = parameters.filter(building__building_number=building_number) 

#         if date:
#             parameters = parameters.filter(created_at=date)

#         if start_date and end_date:
#             parameters = parameters.filter(created_at__range=[start_date, end_date])
        
#         parameters = parameters.order_by('-created_at')

#         serializer = BuildingEnvironmentalParametersSerializer(parameters, many=True)
#         return Response(serializer.data)

#     except Exception as e:
#         print("Error filtering parameters:", str(e))
#         return Response({'error': 'An error occurred while filtering parameters'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def filterBuildingEnvironmentalParameters(request):
    try:
        user = request.user
        logger.info(f"Запрос на фильтрацию параметров окружающей среды для зданий от пользователя: {user.username}")

        responsible_id = request.query_params.get('responsible')
        building_number = request.query_params.get('building')
        date = request.query_params.get('date')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        logger.info(f"Параметры фильтрации: responsible_id={responsible_id}, building_number={building_number}, date={date}, start_date={start_date}, end_date={end_date}")

        parameters = BuildingEnviromentalParameters.objects.all()

        if responsible_id:
            parameters = parameters.filter(responsible__id=responsible_id)

        if building_number:
            parameters = parameters.filter(building__building_number=building_number)

        if date:
            parameters = parameters.filter(created_at=date)

        if start_date and end_date:
            parameters = parameters.filter(created_at__range=[start_date, end_date])

        parameters = parameters.order_by('-created_at')

        serializer = BuildingEnvironmentalParametersSerializer(parameters, many=True)
        logger.info(f"Фильтрация завершена успешно, найдено {parameters.count()} параметров")
        return Response(serializer.data)

    except Exception as e:
        logger.error(f"Произошла ошибка при фильтрации параметров: {e}", exc_info=True)
        return Response({'error': 'Произошла ошибка при фильтрации параметров'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)